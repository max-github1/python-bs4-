from bs4 import BeautifulSoup
import requests
import openpyxl

url_part="https://merolagani.com/CompanyDetail.aspx?symbol="
c_bank=['cbl','kbl','mega']
urls=[]
for i in range(len(c_bank)):
	url=urls.append(url_part+c_bank[i])
for url in urls:
	req=requests.get(url).content
	soup=BeautifulSoup(req,'lxml')
	c_name=soup.find('span', id='ctl00_ContentPlaceHolder1_CompanyDetail1_companyName')
	new_cname=c_name.text
	table=soup.find('table',class_="table table-striped table-hover table-zeromargin", id="accordion")
	new_table=table.text.replace(' ','')

	clean_ntable = ""
	for line in new_table.split("\n"):
	    if line.strip():
	        clean_ntable += line + "\n"


	workbook = openpyxl.Workbook()
	sheet = workbook.active
	sheet.cell(row=1,column=1)
	for i, dat in enumerate(new_table):
		sheet.cell(row=i+2, column=1).value =dat
	workbook.save('tablen.xlsx')
